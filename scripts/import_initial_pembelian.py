from openpyxl import load_workbook
import pdb
import re
from transaction.models import Nasabah, Kategori, Stok, Pembelian

wb = load_workbook('data/manual_value_separated (copy).xlsx')
unregistered_nasabah = []
def extract_price(ws):
	result = []
	for i in range (0, 61):
		old_price = float(ws.cell(row=i+3, column=2).value)
		new_price = float(ws.cell(row=i+3, column=4).value)
		kode = ws.cell(row=i+3, column=7).value
		result.append((kode, old_price, new_price))
	return result

price_dict = extract_price(wb['Update Harga'])

# for i in range(0, len(other_dict)):
# 	if (other_dict[i][0] != price_dict[i][0]) or (other_dict[i][1] - price_dict[i][1] > 0) or (other_dict[i][2] - price_dict[i][2] > 0):
# 		print "DIFF: " + str(other_dict[i]) + " with " + str(price_dict[i])


def explore(ws, needle, max_row, max_column, on_found, default_year, price_category):
	for col in range(1, max_column+1):
		for row in range(1, max_row + 1):
			if ws.cell(row=row, column=col).value == needle:
				on_found(ws, row, col, default_year, price_category)

def extract_nota(ws, initial_row, initial_column):
	for row_step in range(4):
		for col_step in range(6):
			nota = str(ws.cell(row=initial_row + 3 + row_step, column=initial_column + 4 + col_step).value)
			if nota:
				nota_id = re.sub(r'\D', '', nota)
				if len(nota_id) > 4:
					return 'B' + nota_id
	return None

def extract_pembelian(ws, first_row, first_col, default_year, price_category):
	label_offset = 0
	if not ws.cell(row=first_row + 5, column=first_col + 2).value:
		label_offset = -1
	quantity_offset = 0
	if ws.cell(row=first_row + 8, column=first_col + 6).value == 'KODE':
		quantity_offset = 1

	nama_bs = ws.cell(row=first_row + 4, column=first_col + label_offset + 2).value
	tanggal = ws.cell(row=first_row + 5, column=first_col + label_offset + 2).value
	tanggal = tanggal.replace(year=default_year)
	nasabah_id = ws.cell(row=first_row + 6, column=first_col + label_offset + 2).value
	nota = extract_nota(ws, first_row, first_col)
	ref_total = float(ws.cell(row=first_row + 8, column=first_col + quantity_offset + 7).value)
	# Extract quantity
	quantities = []
	for i in range(0, 61):
		current_value = ws.cell(row=first_row + 10 + i, column=first_col + quantity_offset + 6).value
		try:
			current_quantity =  float(current_value)
		except:
			current_quantity = 0
		quantities.append(current_quantity)

	parsed_nasabah_id = None
	try:
		parsed_nasabah_id = int(re.sub(r'\D', '', nasabah_id))
		n = Nasabah.objects.get(id=parsed_nasabah_id)
	except:
		print "Creating new nasabah: " + nama_bs
		n = Nasabah(nama=nama_bs, alamat="-", jenis="kolektif")
		if parsed_nasabah_id:
			n.id = parsed_nasabah_id
		n.save()
		unregistered_nasabah.append((n.id, nama_bs, nota))

	p = Pembelian(tanggal=tanggal, nasabah=n, nota=nota)
	p.save()

	price_index = 1
	if price_category != "old":
		price_index = 2

	current_total = 0
	for i in range(0, 61):
		if quantities[i] == 0:
			continue
		k = Kategori.objects.get(kode=price_dict[i][0])
		s = Stok(kategori = k,
				tanggal = tanggal,
				jumlah = quantities[i],
				harga = price_dict[i][price_index])
		s.save()
		p.stocks.add(s)
		current_total = current_total + quantities[i] * price_dict[i][price_index]
	p.save()
	diff = current_total - ref_total
	if diff != 0:
		mark = ">>>> " + str(diff) + " "
	else:
		mark = ""

	print "{0}[{1}] {2}-{3} on {4}".format(mark, nota, nasabah_id, nama_bs, tanggal)

def explore_worksheet(workbook, ws_name, default_year, price_category):
	FIRST_ENTRY = "Hitungan KG"
	ws1 = workbook[ws_name]
	max_row = len(ws1.columns[0])
	max_column = len(ws1.rows[0])

	explore(ws1, FIRST_ENTRY, max_row, max_column, extract_pembelian, default_year, price_category)

def run():
	explore_worksheet(wb, 'Perhitungan roda 3', 2015, "old")
	explore_worksheet(wb, 'Perhitungan roda 4', 2015, "old")
	explore_worksheet(wb, 'TIMBANG GUDANG', 2015, "old")
	explore_worksheet(wb, 'HAL 2', 2015, "old")
	explore_worksheet(wb, 'HAL 2B', 2015, "old")
	explore_worksheet(wb, 'HAL 2C', 2015, "old")
	explore_worksheet(wb, 'HAL 3', 2016, "new")
	explore_worksheet(wb, 'HAL 3A', 2016, "new")
	explore_worksheet(wb, 'HAL 3B', 2016, "new")
	explore_worksheet(wb, 'HAL 3C', 2016, "new")
	unregistered_nasabah.sort(key= lambda tup: tup[0])
	for new_nasabah in unregistered_nasabah:
		print "{0},{1},{2}".format(new_nasabah[0], new_nasabah[1], new_nasabah[2])
